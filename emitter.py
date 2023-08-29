from reportlab.pdfbase.pdfmetrics import registerFont
from reportlab.pdfbase.ttfonts import TTFont
from threading import active_count, Thread
from reportlab.pdfgen.canvas import Canvas
from pdfrw.toreportlab import makerl
from openpyxl import load_workbook
from time import sleep, time
from textwrap import wrap
from pandas import concat
from math import floor
from reader import *
import neat
import var


def emit_singular(name, cpf, cnpj, matr, clie, apol, capi, cobe='', grup='') -> None:
    """
    Método usado para emitir um certificado. Leva nome, CPF, CNPJ, matrícula, cliente, apólice e coberturas ou código. Retorna nada.
    """

    with var.lock:
        if var.progress == 0:
            # Configura a fonte a ser usada no PDF
            registerFont(TTFont(var.text_font, f'dados\\fontes\\{var.text_font}'))  # Registra essa fonte na memória

        var.progress += 1
        text = var.base_text  # Copia o texto base para uma variável nesse processo.

    # Substitui todos os campos dinâmicos pelos seus respectivos valores nessa instância.
    text = text.replace('<NOME>', neat.name(name))
    text = text.replace('<CPF>', neat.cpf(cpf))
    text = text.replace('<CNPJ>', neat.cnpj(cnpj))
    text = text.replace('<MATRICULA>', neat.matr(matr))
    text = text.replace('<CLIENTE>', neat.name(clie))
    text = text.replace('<APOLICE>', str(apol))
    text = text.replace('<COMECO>', f'{var.start_period.day:02}/{var.start_period.month:02}/{var.start_period.year}')
    text = text.replace('<FINAL>', f'{var.end_period.day:02}/{var.end_period.month:02}/{var.end_period.year}')
    text = text.replace('<CAPITAL>', neat.capi(capi))

    if cobe != '':
        text = text.replace('<COBERTURA>', cobe)

    else:
        text = text.replace('<COBERTURA>', str(grup))

    canvas = Canvas(f'{var.output_dir}{neat.cpf(cpf)} - {neat.name(name)}.pdf')  # Objeto do texto a ser salvo.
    canvas.setFont(var.text_font, var.text_size)  # Seleciona esse fonte para ser usada

    xobj_name = makerl(canvas, var.template_obj)  # Superposição objeto modelo com o objeto escritor por cima.
    canvas.doForm(xobj_name)

    userbreaks = text.split('\n')
    lines = []
    for line in userbreaks:
        broken_lines = wrap(line, var.max_chars)
        for new_line in broken_lines:
            lines.append(new_line)

    y = var.text_height
    for line in lines:
        canvas.drawString(var.dist_left, y, line)  # Escreve a linha por cima do modelo.
        y -= var.line_space  # Espaçamento entre as linhas

    canvas.save()  # Função I/O-intensiva para salvar o arquivo da memória para o armazenamento do computador.


def emit_from_source() -> None:
    """
    Função para emitir diversos certificados advindos de uma fonte de dados externa.
    """

    start_time = time()  # Horário do começo da emissão.
    grupo = get_grupo(var.grupo_dir)

    # Abre a planilha no modo leitura otimizada para obter o número de linhas e então fechá-la.
    row_count_time = time()
    workbook = load_workbook(filename=var.data_dir, read_only=True)
    sheet = workbook.worksheets[0]
    row_count = sheet.max_row - 1  # Desconta os cabeçários
    workbook.close()
    print(f'Tempo necessário para calcular que há {row_count} linhas na planilha: {(time() - row_count_time):.2f} segundos.')

    var.max_progress = row_count
    if row_count > 1000 and var.max_processes != 1:
        # Sabendo o número total de linhas, divide-se o trabalho de carregá-las na memória através de multiprocessamento.
        quotient = floor(var.max_progress / var.max_processes)  # Número de linhas que cada processo irá carregar na memória.
        remainder = var.max_progress % var.max_processes  # Número de linhas a mais que o último processo carregará na memória.
        processes = []

        for i in range(var.max_processes):
            if i == range(var.max_processes)[-1]:  # Se estiver na última iteração, dê a sobra de linhas para o último processo.
                processes.append(SubLoader(var.data_dir, (i * quotient) + 1, quotient + remainder, i, var.shared_list, var.lock))

            elif i == 0:  # Primeira linha. Em todos os casos, deve-se adicionar mais 1, pois senão há overlapping
                processes.append(SubLoader(var.data_dir, 1, quotient, i, var.shared_list, var.lock))

            else:  # Senão, dê apenas o quociente da divisão para esses processos que não são o último.
                processes.append(SubLoader(var.data_dir, (i * quotient) + 1, quotient, i, var.shared_list, var.lock))

            processes[i].start()

        for i in range(var.max_processes):
            processes[i].join()  # Espera todos os processos carregarem as suas seções de dados na memória para então continuar o programa.

        df = concat(var.shared_list[:], ignore_index=True)  # Junta todas as partes num inteiro.
        var.shared_list = var.manager.list(range(var.max_processes))  # Deleta os fragmentos de dados após serem utilizados.

    else:
        df = load(var.data_dir, header=None, skiprows=1)

    threads = []
    emit_time = time()  # Usado para calcular o tempo total e velocidade média do processo de emissão.

    # Usados para calcular velocidade marginal.
    marginal_time = time()
    marginal_cert = 0
    j = 0

    # Isso é necessário para evitar crashes quando o usuário decide continuar mesmo faltando campos.
    # E para caso as palavras-chave estejam vazias.
    arguments = []
    if var.headers['name'] != '':
        arguments.append(lambda i: df.iloc[i, var.headers['name']])
    else:
        arguments.append(lambda i: 'ERRO')

    if var.headers['cpf'] != '':
        arguments.append(lambda i: df.iloc[i, var.headers['cpf']])
    else:
        arguments.append(lambda i: 'ERRO')

    if var.headers['cnpj'] != '':
        arguments.append(lambda i: df.iloc[i, var.headers['cnpj']])
    else:
        arguments.append(lambda i: 'ERRO')

    if var.headers['matr'] != '':
        arguments.append(lambda i: df.iloc[i, var.headers['matr']])
    else:
        arguments.append(lambda i: 'ERRO')

    if var.headers['clie'] != '':
        arguments.append(lambda i: df.iloc[i, var.headers['clie']])
    else:
        arguments.append(lambda i: 'ERRO')

    if var.headers['apol'] != '':
        arguments.append(lambda i: df.iloc[i, var.headers['apol']])
    else:
        arguments.append(lambda i: 'ERRO')

    if var.headers['capi'] != '':
        arguments.append(lambda i: df.iloc[i, var.headers['capi']])
    else:
        arguments.append(lambda i: 'ERRO')

    if var.headers['cobe'] != '' and var.headers['grup'] != '':
        arguments.append(lambda i: df.iloc[i, var.headers['cobe']])
        arguments.append(lambda i: '')
    elif var.headers['cobe'] != '' and var.headers['grup'] == '':
        arguments.append(lambda i: df.iloc[i, var.headers['cobe']])
        arguments.append(lambda i: '')
    elif var.headers['cobe'] == '' and var.headers['grup'] != '':
        arguments.append(lambda i: '')
        arguments.append(lambda i: grupo.loc[neat.grupo(df.iloc[i, var.headers['grup']])])
    else:
        arguments.append(lambda i: 'ERRO')
        arguments.append(lambda i: 'ERRO')

    while j < var.max_progress:
        needed = var.target_threads - active_count()  # Três threads estão sempre ativas: (MainThread, update_progressbar_daemon, e emit_from_source) por isso a opção mínima é quatro.
        if j + needed > var.max_progress:
            needed = var.max_progress - j

        elif needed <= 0:
            needed = 1

        if active_count() < var.max_threads:
            threads.extend(Thread(target=emit_singular, args=[arg(i) for arg in arguments]) for i in range(j, j + needed))

            for i in range(1, needed + 1):
                threads[-i].start()
            j += needed

        else:
            sleep(0.001)

        if marginal_cert != var.progress:
            try:
                print(f'Velocidade marginal: {(var.progress - marginal_cert) / (time() - marginal_time):07.2f} certificados por segundo. | Threads ativas: {active_count():03}. | Progresso: {var.progress:04} de {var.max_progress:04}. | Iniciados: {j:04}. | Threads requeridas: {needed}.')
            except ZeroDivisionError:
                print(f'Velocidade marginal: 9999.99 certificados por segundo. | Threads ativas: {active_count():03}. | Progresso: {var.progress:04} de {var.max_progress:04}. | Iniciados: {j:04}. | Threads requeridas: {needed}.')
            marginal_time = time()
            marginal_cert = var.progress

    for thread in threads:
        thread.join()

    del df

    var.emission_time = time() - start_time
    var.certificates_per_second = var.max_progress / var.emission_time

    var.progress = 0
    print(f'Tempo levado para salvar {var.max_progress} arquivos PDF da memória para o disco: {(time() - emit_time) / 60:.2f} minutos')
    print(f'Velocidade média: {var.certificates_per_second:.2f}')
